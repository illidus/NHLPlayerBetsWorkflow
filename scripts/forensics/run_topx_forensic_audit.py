import argparse
from datetime import datetime, timezone
from pathlib import Path
import os
import re
import sys
from typing import List, Optional, Tuple

import duckdb
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
SRC_ROOT = REPO_ROOT / "src"
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(SRC_ROOT))

from nhl_bets.analysis.normalize import normalize_name
from nhl_bets.analysis.side_integrity import (
    normalize_book,
    normalize_market,
    normalize_player,
    normalize_side,
    resolve_odds_side,
    build_odds_side_lookup,
)
from nhl_bets.projections.config import MARKET_POLICY, get_prob_column_name, ALPHAS
from nhl_bets.projections.single_game_model import compute_game_probs


MARKET_PREFIX = {
    "GOALS": "G",
    "ASSISTS": "A",
    "POINTS": "PTS",
    "SOG": "SOG",
    "BLOCKS": "BLK",
}

MULTIPLIER_KEYS = ["mult_opp_sog", "mult_opp_g", "mult_goalie", "mult_itt", "mult_b2b", "toi_factor"]

MARKET_STAT_FIELDS = {
    "GOALS": ("goals", "Goals"),
    "ASSISTS": ("primary_assists + secondary_assists", "Assists"),
    "POINTS": ("points", "Points"),
    "SOG": ("sog", "Shots on Goal"),
    "BLOCKS": ("blocks", "Blocks"),
}

RECENT_GAME_LIMIT = 10
RECENT_GAME_COLUMNS = [
    col
    for i in range(1, RECENT_GAME_LIMIT + 1)
    for col in (
        f"recentgame{i}_game_id",
        f"recentgame{i}_date",
        f"recentgame{i}_value",
    )
]


def _init_con(db_path: str) -> duckdb.DuckDBPyConnection:
    con = duckdb.connect(db_path)
    con.execute("SET memory_limit = '8GB';")
    con.execute("SET threads = 8;")
    con.execute("SET temp_directory = './duckdb_temp/';")
    return con


def _load_best_bets(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise SystemExit(f"Best bets file not found: {path}")
    df = pd.read_excel(path)
    if df.empty:
        raise SystemExit("Best bets file is empty.")
    
    # Normalize column names from new pipeline format
    rename_map = {
        "EV": "EV%",
        "Imp_Prob": "Implied_Prob"
    }
    df = df.rename(columns=rename_map)

    df["Market"] = df["Market"].astype(str).str.upper()
    df["Side"] = df["Side"].astype(str).str.upper()
    df["Book"] = df["Book"].astype(str)
    df["EV%"] = pd.to_numeric(df["EV%"], errors="coerce")
    return df


def _apply_filters(df: pd.DataFrame, markets: Optional[List[str]], min_ev: Optional[float]) -> pd.DataFrame:
    filtered = df.copy()
    if markets:
        market_set = {m.strip().upper() for m in markets if m.strip()}
        filtered = filtered[filtered["Market"].isin(market_set)].copy()
    if min_ev is not None:
        filtered = filtered[filtered["EV%"] >= min_ev].copy()
    return filtered


def _select_top_x(df: pd.DataFrame, top_x: int) -> pd.DataFrame:
    ranked = df.sort_values("EV%", ascending=False).copy()
    return ranked.head(top_x).reset_index(drop=True)


def _parse_timestamp(value: object):
    return pd.to_datetime(value, errors="coerce", utc=True)


def _resolve_game_date(row: pd.Series):
    if "game_date" in row:
        dt = pd.to_datetime(row.get("game_date"), errors="coerce")
        if pd.notna(dt):
            return dt.normalize()
    ts = _parse_timestamp(row.get("event_start_time_utc"))
    if pd.notna(ts):
        return ts.tz_convert(None).normalize()
    ts = _parse_timestamp(row.get("prob_snapshot_ts"))
    if pd.notna(ts):
        return ts.tz_convert(None).normalize()
    return pd.NaT


def _build_bet_key(row: pd.Series) -> str:
    event_id = row.get("event_id") or row.get("event_id_vendor") or row.get("event_id_raw")
    if pd.notna(event_id):
        return f"event_id:{event_id}"
    game_date = _resolve_game_date(row)
    capture_ts = _parse_timestamp(row.get("capture_ts_utc"))
    parts = [
        f"game_date:{game_date.date() if pd.notna(game_date) else ''}",
        f"player:{str(row.get('Player', '')).strip()}",
        f"team:{str(row.get('Team', '')).strip()}",
        f"opp:{str(row.get('OppTeam', '')).strip()}",
        f"market:{str(row.get('Market', '')).strip()}",
        f"line:{float(row.get('Line')):.3f}" if pd.notna(row.get("Line")) else "line:",
        f"side:{str(row.get('Side', '')).strip()}",
        f"book:{str(row.get('Book', '')).strip()}",
        f"capture_ts:{capture_ts.isoformat() if pd.notna(capture_ts) else ''}",
    ]
    return "|".join(parts)


def _load_players(con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    players = con.execute("SELECT player_id, player_name, team FROM dim_players").df()
    players["norm_name"] = players["player_name"].apply(normalize_name)
    players["team"] = players["team"].astype(str).str.upper()
    return players


def _resolve_player_id(player: str, team: Optional[str], players_df: pd.DataFrame) -> Tuple[Optional[int], str]:
    if not player:
        return None, "MISSING_PLAYER_NAME"
    norm = normalize_name(player)
    matches = players_df[players_df["norm_name"] == norm]
    if matches.empty:
        return None, "PLAYER_NOT_FOUND"
    if team:
        team = str(team).upper()
        team_matches = matches[matches["team"] == team]
        if not team_matches.empty:
            return int(team_matches.iloc[0]["player_id"]), "OK_TEAM_MATCH"
    return int(matches.iloc[0]["player_id"]), "OK_NAME_MATCH"

def _fetch_snapshot_row(
    con: duckdb.DuckDBPyConnection,
    player_id: int,
    game_date: Optional[str] = None,
    game_id: Optional[int] = None,
) -> pd.DataFrame:
    query = f"""
        WITH team_schedule AS (
            SELECT
                game_id,
                home_team AS team,
                game_date,
                season,
                LAG(game_date) OVER (PARTITION BY home_team ORDER BY game_date) AS prev_game_date,
                LAG(game_id) OVER (PARTITION BY home_team ORDER BY game_date) AS prev_game_id
            FROM dim_games
            UNION ALL
            SELECT
                game_id,
                away_team AS team,
                game_date,
                season,
                LAG(game_date) OVER (PARTITION BY away_team ORDER BY game_date) AS prev_game_date,
                LAG(game_id) OVER (PARTITION BY away_team ORDER BY game_date) AS prev_game_id
            FROM dim_games
        ),
        schedule_context AS (
            SELECT
                *,
                CASE WHEN date_diff('day', prev_game_date, game_date) = 1 THEN 1 ELSE 0 END AS is_b2b
            FROM team_schedule
        ),
        recent_roster AS (
            SELECT
                s.game_id,
                s.team,
                s.is_b2b,
                s.prev_game_id,
                gf.goalie_id,
                gf.sum_toi_L10,
                ROW_NUMBER() OVER (PARTITION BY s.game_id, s.team ORDER BY gf.sum_toi_L10 DESC) AS depth_rank
            FROM schedule_context s
            JOIN fact_goalie_features gf
                ON s.team = gf.team
                AND gf.game_date < s.game_date
                AND gf.game_date >= (s.game_date - INTERVAL 14 DAY)
            QUALIFY ROW_NUMBER() OVER (PARTITION BY s.game_id, s.team, gf.goalie_id ORDER BY gf.game_date DESC) = 1
        ),
        prev_starter_info AS (
            SELECT
                s.game_id,
                s.team,
                pg.player_id AS prev_starter_id
            FROM schedule_context s
            JOIN fact_goalie_game_situation pg
                ON s.prev_game_id = pg.game_id
                AND s.team = pg.team
            QUALIFY ROW_NUMBER() OVER (PARTITION BY s.game_id, s.team ORDER BY pg.toi_seconds DESC) = 1
        ),
        primary_goalies AS (
            SELECT
                r.game_id,
                r.team,
                r.goalie_id,
                1 AS rn
            FROM recent_roster r
            LEFT JOIN prev_starter_info p
                ON r.game_id = p.game_id
                AND r.team = p.team
            WHERE
                CASE
                    WHEN r.is_b2b = 1 AND r.depth_rank = 1 AND r.goalie_id = p.prev_starter_id THEN 0
                    WHEN r.is_b2b = 1 AND r.depth_rank = 2 AND (p.prev_starter_id IS NULL OR p.prev_starter_id != r.goalie_id) THEN 1
                    WHEN r.is_b2b = 0 AND r.depth_rank = 1 THEN 1
                    ELSE 0
                END = 1
            QUALIFY ROW_NUMBER() OVER (PARTITION BY r.game_id, r.team ORDER BY r.depth_rank ASC) = 1
        )
        SELECT
            p.player_id,
            p.game_id,
            p.game_date::DATE AS game_date,
            p.season,
            dp.player_name AS Player,
            p.team AS Team,
            p.opp_team AS OppTeam,
            p.home_or_away,
            p.position AS Pos,
            p.xg_per_game_L10 AS G,
            p.goals_per_game_L10 AS G_realized,
            p.assists_per_game_L10 AS A,
            p.points_per_game_L10 AS PTS,
            p.sog_per_game_L10 AS SOG,
            p.blocks_per_game_L10 AS BLK,
            p.ev_ast_60_L20,
            p.pp_ast_60_L20,
            p.ev_pts_60_L20,
            p.pp_pts_60_L20,
            p.ev_toi_minutes_L20,
            p.pp_toi_minutes_L20,
            p.ev_on_ice_xg_60_L20,
            p.pp_on_ice_xg_60_L20,
            p.team_pp_xg_60_L20,
            p.ev_ipp_x_L20,
            p.pp_ipp_x_L20,
            p.primary_ast_ratio_L10,
            p.avg_toi_minutes_L10 AS TOI,
            p.avg_toi_minutes_L10 AS proj_toi,
            d.opp_sa60_L10 AS opp_sa60,
            d.opp_xga60_L10 AS opp_xga60,
            COALESCE(gf.goalie_gsax60_L10, 0.0) AS goalie_gsax60,
            CASE
                WHEN gf.sum_toi_L10 IS NULL OR gf.sum_toi_L10 = 0 THEN 0.0
                ELSE gf.sum_xga_L10 / (gf.sum_toi_L10 / 3600)
            END AS goalie_xga60,
            sc.is_b2b
        FROM fact_player_game_features p
        LEFT JOIN fact_team_defense_features d
            ON p.opp_team = d.team
            AND p.game_date = d.game_date
        LEFT JOIN dim_players dp
            ON p.player_id = dp.player_id
        LEFT JOIN schedule_context sc
            ON p.game_id = sc.game_id
            AND p.team = sc.team
        LEFT JOIN primary_goalies pg
            ON p.game_id = pg.game_id
            AND p.opp_team = pg.team
            AND pg.rn = 1
        LEFT JOIN fact_goalie_features gf
            ON pg.goalie_id = gf.goalie_id
            AND p.game_id = gf.game_id
        WHERE p.player_id = {player_id}
""".strip()

    if game_id is not None:
        query += f"\n          AND p.game_id = {game_id}"
    elif game_date:
        query += f"\n          AND p.game_date::DATE = DATE '{game_date}'"

    return con.execute(query).df()


def _prob_for_line(calcs: dict, market: str, line: float) -> float:
    key = f"probs_{market.lower()}"
    probs = calcs.get(key, {})
    k = int(np.floor(float(line)) + 1)
    return float(probs.get(k, np.nan))


def _calibrated_prob_for_line(calcs: dict, market: str, line: float) -> float:
    if market not in {"ASSISTS", "POINTS"}:
        return np.nan
    k = int(np.floor(float(line)) + 1)
    if market == "ASSISTS":
        probs = calcs.get("probs_assists_calibrated", {})
    else:
        probs = calcs.get("probs_points_calibrated", {})
    return float(probs.get(k, np.nan))


def _production_p_over(market: str, p_over_raw: float, p_over_calibrated: float) -> float:
    policy = MARKET_POLICY.get(market.upper(), "p_over")
    if policy == "p_over_calibrated" and not np.isnan(p_over_calibrated):
        return float(p_over_calibrated)
    return float(p_over_raw)


def _nbinom_inputs(market: str, line: Optional[float], mu_used: float) -> dict:
    inputs = {
        "nbinom_alpha": np.nan,
        "nbinom_k": np.nan,
        "nbinom_n": np.nan,
        "nbinom_p": np.nan,
    }
    alpha = ALPHAS.get(str(market).upper())
    if alpha is None or alpha <= 0 or pd.isna(mu_used) or pd.isna(line):
        return inputs
    try:
        k = int(np.floor(float(line)) + 1)
    except (TypeError, ValueError):
        return inputs
    n = 1.0 / alpha
    p = 1.0 / (1.0 + alpha * float(mu_used))
    inputs.update(
        {
            "nbinom_alpha": alpha,
            "nbinom_k": k,
            "nbinom_n": n,
            "nbinom_p": p,
        }
    )
    return inputs


def _odds_decimal_from_american(odds_american: Optional[float]) -> Optional[float]:
    try:
        odds = float(odds_american)
    except (TypeError, ValueError):
        return None
    if odds == 0:
        return None
    if odds > 0:
        return 1.0 + (odds / 100.0)
    return 1.0 + (100.0 / abs(odds))

def _match_prob_snapshot(row: pd.Series, prob_df: pd.DataFrame) -> Optional[pd.Series]:
    game_date = _resolve_game_date(row)
    if pd.isna(game_date):
        return None
    player_key = normalize_name(row.get("Player", ""))
    subset = prob_df[
        (prob_df["player_key"] == player_key) & (prob_df["game_date"] == game_date.date())
    ]
    if subset.empty:
        return None
    team = row.get("Team")
    if pd.notna(team):
        team = str(team).upper()
        team_subset = subset[subset["Team"].str.upper() == team]
        if not team_subset.empty:
            return team_subset.iloc[0]
    return subset.iloc[0]


def _build_projection_trace(
    row: pd.Series,
    con: duckdb.DuckDBPyConnection,
    players_df: pd.DataFrame,
    base_proj_df: Optional[pd.DataFrame] = None,
    prob_snapshot_df: Optional[pd.DataFrame] = None,
) -> dict:
    canonical_game_id = row.get("canonical_game_id")
    try:
        canonical_game_id = int(canonical_game_id) if not pd.isna(canonical_game_id) else None
    except (TypeError, ValueError):
        canonical_game_id = None

    result = {
        "projection_status": "OK",
        "projection_note": "",
        "LINE_MATCH": False,
        "PROB_MATH_MATCH": False,
        "CALIBRATION_PLATEAU_EFFECT": False,
        "MULTIPLIER_OUTLIER": False,
        "MU_IMPLAUSIBLE": False,
        "p_over_raw": np.nan,
        "p_over_calibrated": np.nan,
        "p_over_prod": np.nan,
        "p_used": np.nan,
        "ev_calc": np.nan,
        "mu_used": np.nan,
        "multipliers": {},
        "snapshot_inputs": {},
        "recent_stat_label": "",
        "recent_stat_values": [],
        "base_projection": {},
        "snapshot_game_id": None,
        "canonical_game_id": canonical_game_id,
        "event_id_vendor": row.get("event_id_vendor"),
        "event_start_time_utc": row.get("event_start_time_utc"),
        "source_capture_ts": row.get("capture_ts_utc"),
        "source_vendor": row.get("source_vendor"),
    }
    market = str(row.get("Market", "")).upper()
    line = row.get("Line")
    side = normalize_side(row.get("Side"))
    game_date = _resolve_game_date(row)
    if pd.isna(game_date) and canonical_game_id is None:
        result.update(
            {
                "projection_status": "MISSING_GAME_DATE",
                "projection_note": "No game date available for snapshot lookup.",
            }
        )
        return result

    player_id, player_note = _resolve_player_id(row.get("Player"), row.get("Team"), players_df)
    result["player_id"] = player_id
    result["player_id_note"] = player_note
    base_row = None
    if base_proj_df is not None:
        base_row = _match_base_projection(base_proj_df, row.get("Player"), row.get("Team"))
    if base_row is not None:
        result["base_projection"] = base_row.to_dict()
    stat_info = MARKET_STAT_FIELDS.get(market)
    stat_label = ""
    recent_stats = []
    if player_id and stat_info:
        stat_label = stat_info[1]
        recent_stats = _stat_recent_values(con, player_id, stat_info[0])
    result["recent_stat_label"] = stat_label
    result["recent_stat_values"] = recent_stats
    if player_id is None:
        result.update(
            {
                "projection_status": "PLAYER_ID_NOT_FOUND",
                "projection_note": player_note,
            }
        )
        return result

    game_date_str = game_date.date().isoformat() if pd.notna(game_date) else None
    snapshot_df = _fetch_snapshot_row(con, player_id, game_date_str, canonical_game_id)
    if snapshot_df.empty:
        fallback_row = None
        if prob_snapshot_df is not None:
            fallback_row = _match_prob_snapshot(row, prob_snapshot_df)
        if fallback_row is None:
            market = str(row.get("Market", "")).upper()
            line = row.get("Line")
            p_over_raw = row.get("p_over_raw")
            p_over_calibrated = row.get("p_over_calibrated")
            p_over_prod = row.get("p_over_selected")
            if pd.isna(p_over_prod) and pd.notna(p_over_raw):
                p_over_prod = _production_p_over(market, p_over_raw, p_over_calibrated)

            if side == "OVER":
                p_used = p_over_prod if pd.notna(p_over_prod) else np.nan
            elif side == "UNDER":
                p_used = 1.0 - p_over_prod if pd.notna(p_over_prod) else np.nan
            else:
                p_used = np.nan

            odds_decimal = row.get("odds_decimal")
            if odds_decimal is None or pd.isna(odds_decimal):
                odds_decimal = _odds_decimal_from_american(row.get("Odds"))
            ev_calc = (p_used * odds_decimal - 1.0) if odds_decimal else np.nan

            mu_used = row.get("mu_adj_value")
            if pd.isna(mu_used):
                mu_used = row.get("mu")

            tol = 1e-6
            line_match = pd.notna(p_over_raw)
            calib_plateau = (
                market in {"ASSISTS", "POINTS"}
                and float(line) == 0.5
                and pd.notna(p_over_calibrated)
                and pd.notna(p_over_raw)
                and abs(float(p_over_calibrated) - float(p_over_raw)) > tol
            )
            mu_implausible = pd.notna(mu_used) and (float(mu_used) <= 0.0 or float(mu_used) > 6.0)

            result.update(
                {
                    "projection_status": "OK_BEST_BETS",
                    "projection_note": "Used best-bets projection fields; snapshot unavailable.",
                    "game_date": game_date.date().isoformat(),
                    "p_over_raw": p_over_raw,
                    "p_over_calibrated": p_over_calibrated,
                    "p_over_prod": p_over_prod,
                    "p_used": p_used,
                    "ev_calc": ev_calc,
                    "mu_used": mu_used,
                    "LINE_MATCH": line_match,
                    "PROB_MATH_MATCH": True,
                    "CALIBRATION_PLATEAU_EFFECT": calib_plateau,
                    "MULTIPLIER_OUTLIER": False,
                    "MU_IMPLAUSIBLE": mu_implausible,
                    "multipliers": {},
                    "snapshot_inputs": {},
                }
            )
            return result

        market = str(row.get("Market", "")).upper()
        line = row.get("Line")
        raw_col = get_prob_column_name(market.lower(), line, "p_over")
        calib_col = get_prob_column_name(market.lower(), line, "p_over_calibrated")
        p_over_raw = float(fallback_row.get(raw_col, np.nan))
        p_over_calibrated = float(fallback_row.get(calib_col, np.nan))
        p_over_prod = _production_p_over(market, p_over_raw, p_over_calibrated)
        if side == "OVER":
            p_used = p_over_prod
        elif side == "UNDER":
            p_used = 1.0 - p_over_prod if pd.notna(p_over_prod) else np.nan
        else:
            p_used = np.nan

        odds_decimal = row.get("odds_decimal")
        if odds_decimal is None or pd.isna(odds_decimal):
            odds_decimal = _odds_decimal_from_american(row.get("Odds"))
        ev_calc = (p_used * odds_decimal - 1.0) if odds_decimal else np.nan

        mu_col_map = {
            "GOALS": "mu_adj_G",
            "ASSISTS": "mu_adj_A",
            "POINTS": "mu_adj_PTS",
            "SOG": "mu_adj_SOG",
            "BLOCKS": "mu_adj_BLK",
        }
        mu_used = fallback_row.get(mu_col_map.get(market, ""), np.nan)

        best_raw = row.get("p_over_raw")
        best_calib = row.get("p_over_calibrated")
        best_selected = row.get("p_over_selected")
        tol = 1e-6
        prob_match = False
        if pd.notna(best_raw) and pd.notna(p_over_raw):
            prob_match = abs(float(best_raw) - float(p_over_raw)) <= tol
        if not prob_match and pd.notna(best_calib) and pd.notna(p_over_calibrated):
            prob_match = abs(float(best_calib) - float(p_over_calibrated)) <= tol
        if not prob_match and pd.notna(best_selected) and pd.notna(p_over_prod):
            prob_match = abs(float(best_selected) - float(p_over_prod)) <= tol

        line_match = not np.isnan(p_over_raw)
        calib_plateau = (
            market in {"ASSISTS", "POINTS"}
            and float(line) == 0.5
            and pd.notna(p_over_calibrated)
            and pd.notna(p_over_raw)
            and abs(float(p_over_calibrated) - float(p_over_raw)) > tol
        )
        multipliers = {
            "mult_opp_sog": fallback_row.get("mult_opp_sog"),
            "mult_opp_g": fallback_row.get("mult_opp_g"),
            "mult_goalie": fallback_row.get("mult_goalie"),
            "mult_itt": fallback_row.get("mult_itt"),
            "mult_b2b": fallback_row.get("mult_b2b"),
        }
        multiplier_outlier = any(
            pd.notna(multipliers.get(key))
            and (float(multipliers.get(key)) < 0.5 or float(multipliers.get(key)) > 1.5)
            for key in multipliers
        )
        mu_implausible = pd.notna(mu_used) and (float(mu_used) <= 0.0 or float(mu_used) > 6.0)

        result.update(
            {
                "projection_status": "OK_FALLBACK",
                "projection_note": "Used SingleGamePropProbabilities.csv snapshot.",
                "game_date": game_date.date().isoformat(),
                "p_over_raw": p_over_raw,
                "p_over_calibrated": p_over_calibrated,
                "p_over_prod": p_over_prod,
                "p_used": p_used,
                "ev_calc": ev_calc,
                "mu_used": mu_used,
                "LINE_MATCH": line_match,
                "PROB_MATH_MATCH": prob_match,
                "CALIBRATION_PLATEAU_EFFECT": calib_plateau,
                "MULTIPLIER_OUTLIER": multiplier_outlier,
                "MU_IMPLAUSIBLE": mu_implausible,
                "multipliers": multipliers,
                "snapshot_inputs": fallback_row.to_dict(),
            }
        )
        return result

    snapshot = snapshot_df.iloc[0].to_dict()
    result["snapshot_game_id"] = snapshot.get("game_id")
    context = {
        "opp_sa60": snapshot.get("opp_sa60"),
        "opp_xga60": snapshot.get("opp_xga60"),
        "goalie_gsax60": snapshot.get("goalie_gsax60"),
        "goalie_xga60": snapshot.get("goalie_xga60"),
        "implied_team_total": None,
        "is_b2b": snapshot.get("is_b2b"),
        "proj_toi": snapshot.get("proj_toi"),
    }
    calcs = compute_game_probs(snapshot, context)

    p_over_raw = _prob_for_line(calcs, market, line)
    p_over_calibrated = _calibrated_prob_for_line(calcs, market, line)
    p_over_prod = _production_p_over(market, p_over_raw, p_over_calibrated)

    if side == "OVER":
        p_used = p_over_prod if pd.notna(p_over_prod) else np.nan
    elif side == "UNDER":
        p_used = 1.0 - p_over_prod if pd.notna(p_over_prod) else np.nan
    else:
        p_used = np.nan

    odds_decimal = row.get("odds_decimal")
    if odds_decimal is None or pd.isna(odds_decimal):
        odds_decimal = _odds_decimal_from_american(row.get("Odds"))
    ev_calc = (p_used * odds_decimal - 1.0) if odds_decimal else np.nan

    mu_map = {
        "GOALS": calcs.get("mu_goals"),
        "ASSISTS": calcs.get("mu_assists"),
        "POINTS": calcs.get("mu_points"),
        "SOG": calcs.get("mu_sog"),
        "BLOCKS": calcs.get("mu_blocks"),
    }
    mu_used = mu_map.get(market, np.nan)

    best_raw = row.get("p_over_raw")
    best_calib = row.get("p_over_calibrated")
    best_selected = row.get("p_over_selected")
    tol = 1e-6
    prob_match = False
    if pd.notna(best_raw) and pd.notna(p_over_raw):
        prob_match = abs(float(best_raw) - float(p_over_raw)) <= tol
    if not prob_match and pd.notna(best_calib) and pd.notna(p_over_calibrated):
        prob_match = abs(float(best_calib) - float(p_over_calibrated)) <= tol
    if not prob_match and pd.notna(best_selected) and pd.notna(p_over_prod):
        prob_match = abs(float(best_selected) - float(p_over_prod)) <= tol

    line_match = not np.isnan(p_over_raw)
    calib_plateau = (
        market in {"ASSISTS", "POINTS"}
        and float(line) == 0.5
        and pd.notna(p_over_calibrated)
        and pd.notna(p_over_raw)
        and abs(float(p_over_calibrated) - float(p_over_raw)) > tol
    )
    multiplier_outlier = any(
        pd.notna(calcs.get(key))
        and (float(calcs.get(key)) < 0.5 or float(calcs.get(key)) > 1.5)
        for key in MULTIPLIER_KEYS
    )
    mu_implausible = pd.notna(mu_used) and (float(mu_used) <= 0.0 or float(mu_used) > 6.0)

    result.update(
        {
            "projection_status": "OK",
            "projection_note": "",
            "game_date": game_date.date().isoformat(),
            "p_over_raw": p_over_raw,
            "p_over_calibrated": p_over_calibrated,
            "p_over_prod": p_over_prod,
            "p_used": p_used,
            "ev_calc": ev_calc,
            "mu_used": mu_used,
            "LINE_MATCH": line_match,
            "PROB_MATH_MATCH": prob_match,
            "CALIBRATION_PLATEAU_EFFECT": calib_plateau,
            "MULTIPLIER_OUTLIER": multiplier_outlier,
            "MU_IMPLAUSIBLE": mu_implausible,
            "multipliers": {key: calcs.get(key) for key in MULTIPLIER_KEYS},
            "snapshot_inputs": snapshot,
        }
    )
    return result

def _load_odds(con: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    query_lines = [
        "SELECT",
        "    source_vendor,",
        "    capture_ts_utc,",
        "    event_start_time_utc,",
        "    event_id_vendor,",
        "    player_name_raw,",
        "    market_type,",
        "    line,",
        "    side,",
        "    book_name_raw,",
        "    odds_american,",
        "    odds_decimal,",
        "    raw_payload_path,",
        "    raw_payload_hash",
        "FROM fact_prop_odds",
    ]
    query = "\n".join(query_lines)
    df = con.execute(query).df()
    if df.empty:
        return df
    df["event_date"] = pd.to_datetime(df["event_start_time_utc"], errors="coerce").dt.date
    df["player_key"] = df["player_name_raw"].apply(normalize_player)
    df["market_key"] = df["market_type"].apply(normalize_market)
    df["line_key"] = pd.to_numeric(df["line"], errors="coerce").round(3)
    df["book_key"] = df["book_name_raw"].apply(normalize_book)
    return df


def _select_candidates(trace_row: pd.Series, odds_df: pd.DataFrame) -> pd.DataFrame:
    mask = (
        (odds_df["player_key"] == trace_row["player_key"])
        & (odds_df["market_key"] == trace_row["market_key"])
        & (odds_df["line_key"] == trace_row["line_key"])
        & (odds_df["book_key"] == trace_row["book_key"])
    )
    return odds_df[mask].copy()


def _apply_date_preference(trace_row: pd.Series, candidates: pd.DataFrame) -> pd.DataFrame:
    if candidates.empty:
        return candidates
    game_date = trace_row.get("game_date_key")
    if game_date and candidates["event_date"].notna().any():
        date_match = candidates[candidates["event_date"] == game_date]
        if not date_match.empty:
            candidates = date_match
        else:
            candidates = candidates.copy()
            def _compute_delta(d):
                if pd.isna(d):
                    return None
                try:
                    d_date = pd.to_datetime(d).date()
                except Exception:
                    return None
                if pd.isna(game_date):
                    return None
                return abs((d_date - game_date).days)
            candidates["date_delta"] = candidates["event_date"].apply(_compute_delta)
            min_delta = candidates["date_delta"].min()
            if pd.notna(min_delta):
                candidates = candidates[candidates["date_delta"] == min_delta]
    return candidates


def _apply_odds_preference(trace_row: pd.Series, candidates: pd.DataFrame) -> pd.DataFrame:
    if candidates.empty:
        return candidates
    target_odds = trace_row.get("odds_american")
    if pd.notna(target_odds):
        matches = candidates[candidates["odds_american"] == int(target_odds)]
        if not matches.empty:
            return matches
    return candidates


def _apply_asof_preference(trace_row: pd.Series, candidates: pd.DataFrame) -> pd.DataFrame:
    if candidates.empty:
        return candidates
    asof_ts = trace_row.get("capture_ts_utc")
    if pd.isna(asof_ts):
        return candidates
    candidates["capture_ts_utc"] = pd.to_datetime(candidates["capture_ts_utc"], errors="coerce", utc=True)
    asof_ts = pd.to_datetime(asof_ts, errors="coerce", utc=True)
    if pd.isna(asof_ts):
        return candidates
    filtered = candidates[candidates["capture_ts_utc"] <= asof_ts]
    if filtered.empty:
        return candidates
    max_ts = filtered["capture_ts_utc"].max()
    return filtered[filtered["capture_ts_utc"] == max_ts]


def _build_trace_keys(df: pd.DataFrame) -> pd.DataFrame:
    keys = df.copy()
    keys["player_key"] = keys["Player"].apply(normalize_player)
    keys["market_key"] = keys["Market"].apply(normalize_market)
    keys["line_key"] = pd.to_numeric(keys["Line"], errors="coerce").round(3)
    keys["book_key"] = keys["Book"].apply(normalize_book)
    keys["bet_side_norm"] = keys["Side"].apply(normalize_side)
    keys["game_date_key"] = keys.apply(lambda r: _resolve_game_date(r), axis=1).dt.date
    keys["capture_ts_utc"] = pd.to_datetime(keys.get("capture_ts_utc"), errors="coerce", utc=True)
    return keys


def _load_base_projections(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_csv(path)
    return df


def _match_base_projection(
    base_df: pd.DataFrame, player: str, team: str | None
) -> Optional[pd.Series]:
    if base_df.empty:
        return None
    subset = base_df[base_df["Player"] == player]
    if team:
        subset = subset[subset["Team"].str.upper() == str(team).upper()]
    if subset.empty:
        return None
    if "SOG Per Game" in subset.columns:
        subset = subset.sort_values("SOG Per Game", ascending=False)
    return subset.iloc[0]


def _stat_recent_values(
    con: duckdb.DuckDBPyConnection, player_id: int, stat_expr: str, limit: int = RECENT_GAME_LIMIT
) -> list[dict]:
    query_lines = [
        "SELECT game_id, game_date,",
        f"       {stat_expr} AS stat_value",
        "FROM fact_skater_game_situation",
        "WHERE player_id = ?",
        "  AND situation = 'all'",
        f"  AND {stat_expr} IS NOT NULL",
        "ORDER BY game_date DESC",
        f"LIMIT {limit}",
    ]
    query = "\n".join(query_lines)
    df = con.execute(query, [player_id]).df()
    result = []
    for row in df.itertuples(index=False):
        result.append(
            {
                "game_date": row.game_date.date().isoformat()
                if hasattr(row.game_date, "date")
                else str(row.game_date),
                "game_id": getattr(row, "game_id", None),
                "value": float(row.stat_value),
            }
        )
    return result

def _recent_stats_average(recent_stats: list[dict]) -> float:
    if not recent_stats:
        return np.nan
    return float(sum(entry["value"] for entry in recent_stats) / len(recent_stats))

def _run_odds_audit(
    trace_row: pd.Series,
    odds_df: pd.DataFrame,
    side_lookup: dict,
    use_asof: bool,
) -> dict:
    candidates = _select_candidates(trace_row, odds_df)
    candidate_count = len(candidates)
    candidates = _apply_date_preference(trace_row, candidates)
    candidates = _apply_odds_preference(trace_row, candidates)
    if use_asof:
        candidates = _apply_asof_preference(trace_row, candidates)

    if candidates.empty:
        return {
            "ODDS_MATCHED": False,
            "SIDE_MATCH": False,
            "AMBIGUOUS_MATCH_COUNT": 0,
            "odds_candidates": [],
            "odds_selected": None,
        }

    candidates = candidates.copy()
    candidates["capture_ts_utc"] = pd.to_datetime(candidates["capture_ts_utc"], errors="coerce", utc=True)
    candidates = candidates.sort_values("capture_ts_utc", ascending=False)
    selected = candidates.iloc[0]

    ambiguous_count = 0
    candidate_rows = []
    for _, cand in candidates.iterrows():
        odds_side_interpreted = normalize_side(cand["side"])
        odds_side_lookup, lookup_reason = resolve_odds_side(
            side_lookup,
            cand["player_name_raw"],
            cand["market_type"],
            cand["line"],
            cand["book_name_raw"],
            cand["odds_american"],
        )
        if odds_side_lookup == "AMBIGUOUS":
            ambiguous_count += 1
        candidate_rows.append(
            {
                "odds_source_vendor": cand["source_vendor"],
                "odds_capture_ts_utc": cand["capture_ts_utc"],
                "odds_event_start_time_utc": cand["event_start_time_utc"],
                "odds_event_date": cand["event_date"],
                "odds_player_raw": cand["player_name_raw"],
                "odds_market_raw": cand["market_type"],
                "odds_line": cand["line"],
                "odds_side_raw": cand["side"],
                "odds_side_interpreted": odds_side_interpreted,
                "odds_side_lookup": odds_side_lookup,
                "odds_side_lookup_reason": lookup_reason,
                "odds_book_raw": cand["book_name_raw"],
                "odds_american": cand["odds_american"],
                "odds_decimal": cand["odds_decimal"],
                "odds_raw_payload_path": cand["raw_payload_path"],
                "odds_raw_payload_hash": cand["raw_payload_hash"],
            }
        )

    odds_side_interpreted = normalize_side(selected["side"])
    bet_side = trace_row.get("bet_side_norm")
    side_match = bet_side == odds_side_interpreted if bet_side else False

    odds_selected = candidate_rows[0] if candidate_rows else None
    odds_selected["candidate_count"] = candidate_count

    return {
        "ODDS_MATCHED": True,
        "SIDE_MATCH": side_match,
        "AMBIGUOUS_MATCH_COUNT": ambiguous_count,
        "odds_candidates": candidate_rows,
        "odds_selected": odds_selected,
    }


def _classify_row(projection: dict, odds_audit: dict) -> str:
    if projection.get("projection_status") not in {"OK", "OK_FALLBACK", "OK_BEST_BETS"}:
        return "PROJECTION_INPUT/MULTIPLIER ISSUE"
    projection_issue = (
        not projection.get("PROB_MATH_MATCH", False)
        or projection.get("MULTIPLIER_OUTLIER", False)
        or projection.get("MU_IMPLAUSIBLE", False)
    )
    if projection_issue:
        return "PROJECTION_INPUT/MULTIPLIER ISSUE"
    if projection.get("CALIBRATION_PLATEAU_EFFECT", False):
        return "CALIBRATION_PLATEAU_EFFECT"
    odds_issue = (
        not odds_audit.get("ODDS_MATCHED", False)
        or not odds_audit.get("SIDE_MATCH", False)
        or odds_audit.get("AMBIGUOUS_MATCH_COUNT", 0) > 0
    )
    if odds_issue:
        return "ODDS JOIN / SIDE ISSUE"
    return "LEGITIMATE"


def _format_markdown_table(df: pd.DataFrame, max_rows: int = 20) -> str:
    if df.empty:
        return "No rows.\n"
    return df.head(max_rows).to_markdown(index=False) + "\n"


def _coerce_excel_value(value):
    if isinstance(value, pd.Timestamp):
        if value.tzinfo is not None:
            return value.tz_convert(None)
        return value.to_pydatetime()
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _write_formula_workbook(path: Path, df: pd.DataFrame) -> None:
    target_cols = [
        "ev_recomputed",
        "p_used",
        "p_over_selected",
        "p_over_raw",
        "p_over_calibrated",
        "mu_used",
        "recent_stat_avg",
        "nbinom_n",
        "nbinom_p",
        "base_mu_input",
    ]
    base_columns = df.columns.tolist()
    column_order = []
    for col in base_columns:
        column_order.append(col)
        if col in target_cols:
            column_order.append(f"{col}_formula")
            column_order.append(f"{col}_delta")

    wb = Workbook()
    ws = wb.active
    col_index_map = {col: idx + 1 for idx, col in enumerate(column_order)}
    
    for col, idx in col_index_map.items():
        ws.cell(row=1, column=idx, value=col)
        if col.endswith("_formula"):
            label = "formula"
        elif col.endswith("_delta"):
            label = "delta"
        else:
            label = "value"
        ws.cell(row=2, column=idx, value=label)

    def _get_col_letter(col_name):
        idx = col_index_map.get(col_name)
        return get_column_letter(idx) if idx else None

    letters = {}
    for key in [
        "side", "market", "odds_decimal", "base_mu_input",
        "mult_opp_sog", "mult_opp_g", "mult_goalie", "mult_itt", "mult_b2b", "toi_factor",
        "nbinom_k", "nbinom_n", "nbinom_p", "nbinom_alpha", "mu_used",
    ]:
        letters[key] = _get_col_letter(key)
    
    for t in target_cols:
        letters[t] = _get_col_letter(t)
        letters[f"{t}_formula"] = _get_col_letter(f"{t}_formula")
        letters[f"{t}_delta"] = _get_col_letter(f"{t}_delta")
    
    recent_game_letters = []
    for i in range(1, 11):
        l = _get_col_letter(f"recentgame{i}_value")
        if l:
            recent_game_letters.append(l)

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 3
        for col in base_columns:
            ws.cell(row=excel_row, column=col_index_map[col], value=_coerce_excel_value(row[col]))

        # 1. p_used
        if letters.get("p_used_formula") and letters.get("side") and letters.get("p_over_selected"):
            formula = f"=IF({letters['side']}{excel_row}=\"UNDER\",1-{letters['p_over_selected']}{excel_row},{letters['p_over_selected']}{excel_row})"
            ws.cell(row=excel_row, column=col_index_map["p_used_formula"], value=formula)
            if letters.get("p_used_delta"):
                ws.cell(row=excel_row, column=col_index_map["p_used_delta"], 
                        value=f"={letters['p_used_formula']}{excel_row}-{letters['p_used']}{excel_row}")

        # 2. ev_recomputed
        if letters.get("ev_recomputed_formula") and letters.get("p_used") and letters.get("odds_decimal"):
            formula = f"=({letters['p_used']}{excel_row}*{letters['odds_decimal']}{excel_row})-1"
            ws.cell(row=excel_row, column=col_index_map["ev_recomputed_formula"], value=formula)
            if letters.get("ev_recomputed_delta"):
                ws.cell(row=excel_row, column=col_index_map["ev_recomputed_delta"], 
                        value=f"={letters['ev_recomputed_formula']}{excel_row}-{letters['ev_recomputed']}{excel_row}")

        # 3. p_over_selected
        if letters.get("p_over_selected_formula") and letters.get("market") and letters.get("p_over_calibrated") and letters.get("p_over_raw"):
            formula = f"=IF(OR({letters['market']}{excel_row}=\"ASSISTS\",{letters['market']}{excel_row}=\"POINTS\"),{letters['p_over_calibrated']}{excel_row},{letters['p_over_raw']}{excel_row})"
            ws.cell(row=excel_row, column=col_index_map["p_over_selected_formula"], value=formula)
            if letters.get("p_over_selected_delta"):
                ws.cell(row=excel_row, column=col_index_map["p_over_selected_delta"], 
                        value=f"={letters['p_over_selected_formula']}{excel_row}-{letters['p_over_selected']}{excel_row}")

        # 4. p_over_raw: Manual Gamma Summation for float-n support
        if letters.get("p_over_raw_formula") and letters.get("nbinom_k") and letters.get("nbinom_n") and letters.get("nbinom_p"):
            try:
                k_val = row.get("nbinom_k")
                if pd.notna(k_val):
                    k_int = int(k_val)
                    # CDF(k-1), so sum from i=0 to k-1
                    limit = k_int - 1
                    terms = []
                    n_ref = f"{letters['nbinom_n']}{excel_row}"
                    p_ref = f"{letters['nbinom_p']}{excel_row}"
                    
                    if limit >= 0:
                        for i in range(limit + 1):
                            # Excel: EXP(GAMMALN(n+i) - GAMMALN(n) - GAMMALN(i+1)) * p^n * (1-p)^i
                            term = (
                                f"(EXP(GAMMALN({n_ref}+{i})-GAMMALN({n_ref})-GAMMALN({i+1}))"
                                f"*POWER({p_ref},{n_ref})*POWER(1-{p_ref},{i}))"
                            )
                            terms.append(term)
                        cdf_expr = "+".join(terms)
                        formula = f"=1-({cdf_expr})"
                    else:
                        formula = "=1" 
                        
                    ws.cell(row=excel_row, column=col_index_map["p_over_raw_formula"], value=formula)
                else:
                     ws.cell(row=excel_row, column=col_index_map["p_over_raw_formula"], value="")
            except Exception:
                 ws.cell(row=excel_row, column=col_index_map["p_over_raw_formula"], value="ERROR")

            if letters.get("p_over_raw_delta"):
                val_ref = f"{letters['p_over_raw']}{excel_row}"
                form_ref = f"{letters['p_over_raw_formula']}{excel_row}"
                ws.cell(row=excel_row, column=col_index_map["p_over_raw_delta"], 
                        value=f"=IF(ISNUMBER({form_ref}),{form_ref}-{val_ref},\"\")")

        # 5. mu_used
        if letters.get("mu_used_formula") and letters.get("base_mu_input"):
            mult_cols = ["mult_opp_sog", "mult_opp_g", "mult_goalie", "mult_itt", "mult_b2b", "toi_factor"]
            factors = [f"{letters[m]}{excel_row}" for m in mult_cols if letters.get(m)]
            
            # Special logic for SOG: Double application of TOI factor due to model implementation
            # (Once in base_mu calculation implicitly via proj_toi, once explicitly in adjustment)
            # base_mu_input for SOG is calculated as (Corsi * Thru * BaseTOI/60)
            # mu_used (Python) = (Corsi * Thru * ProjTOI/60) * Multipliers * TOI_Factor
            #                  = (Corsi * Thru * BaseTOI * TOI_Factor / 60) * Multipliers * TOI_Factor
            #                  = base_mu_input * TOI_Factor * Multipliers * TOI_Factor
            
            market_val = row.get("market", "")
            is_sog = str(market_val).upper() == "SOG"
            
            if is_sog and letters.get("toi_factor"):
                # Append toi_factor again to the factors list
                factors.append(f"{letters['toi_factor']}{excel_row}")

            if factors:
                formula = f"={letters['base_mu_input']}{excel_row}*{'*'.join(factors)}"
                ws.cell(row=excel_row, column=col_index_map["mu_used_formula"], value=formula)
                if letters.get("mu_used_delta"):
                    ws.cell(row=excel_row, column=col_index_map["mu_used_delta"], 
                            value=f"={letters['mu_used_formula']}{excel_row}-{letters['mu_used']}{excel_row}")

        # 6. recent_stat_avg
        if letters.get("recent_stat_avg_formula") and recent_game_letters:
            formula = f"=AVERAGE({','.join([f'{l}{excel_row}' for l in recent_game_letters])})"
            ws.cell(row=excel_row, column=col_index_map["recent_stat_avg_formula"], value=formula)
            if letters.get("recent_stat_avg_delta"):
                ws.cell(row=excel_row, column=col_index_map["recent_stat_avg_delta"], 
                        value=f"={letters['recent_stat_avg_formula']}{excel_row}-{letters['recent_stat_avg']}{excel_row}")

        # 7. p_over_calibrated (Identity)
        if letters.get("p_over_calibrated_formula") and letters.get("p_over_calibrated"):
             ws.cell(row=excel_row, column=col_index_map["p_over_calibrated_formula"], value=f"={letters['p_over_calibrated']}{excel_row}")
             if letters.get("p_over_calibrated_delta"):
                 ws.cell(row=excel_row, column=col_index_map["p_over_calibrated_delta"], 
                         value=f"={letters['p_over_calibrated_formula']}{excel_row}-{letters['p_over_calibrated']}{excel_row}")

        # 8. nbinom_n: =1/nbinom_alpha
        if letters.get("nbinom_n_formula") and letters.get("nbinom_alpha"):
            formula = f"=1/{letters['nbinom_alpha']}{excel_row}"
            ws.cell(row=excel_row, column=col_index_map["nbinom_n_formula"], value=formula)
            if letters.get("nbinom_n_delta"):
                ws.cell(row=excel_row, column=col_index_map["nbinom_n_delta"], 
                        value=f"={letters['nbinom_n_formula']}{excel_row}-{letters['nbinom_n']}{excel_row}")

        # 9. nbinom_p: =1/(1 + nbinom_alpha * mu_used)
        if letters.get("nbinom_p_formula") and letters.get("nbinom_alpha") and letters.get("mu_used"):
            formula = f"=1/(1+{letters['nbinom_alpha']}{excel_row}*{letters['mu_used']}{excel_row})"
            ws.cell(row=excel_row, column=col_index_map["nbinom_p_formula"], value=formula)
            if letters.get("nbinom_p_delta"):
                ws.cell(row=excel_row, column=col_index_map["nbinom_p_delta"], 
                        value=f"={letters['nbinom_p_formula']}{excel_row}-{letters['nbinom_p']}{excel_row}")

        # 10. base_mu_input: =AVERAGE(recentgame1..10)
        if letters.get("base_mu_input_formula") and recent_game_letters:
            formula = f"=AVERAGE({','.join([f'{l}{excel_row}' for l in recent_game_letters])})"
            ws.cell(row=excel_row, column=col_index_map["base_mu_input_formula"], value=formula)
            if letters.get("base_mu_input_delta"):
                ws.cell(row=excel_row, column=col_index_map["base_mu_input_delta"], 
                        value=f"={letters['base_mu_input_formula']}{excel_row}-{letters['base_mu_input']}{excel_row}")

    wb.save(path)

def main() -> None:
    parser = argparse.ArgumentParser(description="Top-X forensic audit (projection + odds side integrity).")
    parser.add_argument("--top-x", type=int, default=50)
    parser.add_argument("--source-xlsx", default="outputs/ev_analysis/ev_bets_ranked.xlsx")
    parser.add_argument("--outdir", default=None)
    parser.add_argument("--markets", default=None, help="Comma-separated markets filter (e.g. GOALS,ASSISTS).")
    parser.add_argument("--min-ev", type=float, default=None, help="Minimum EV% threshold (e.g. 0.05).")
    parser.add_argument("--side-integrity-guard", action="store_true")
    parser.add_argument("--odds-asof-join", action="store_true")
    parser.add_argument("--duckdb-path", default="data/db/nhl_backtest.duckdb")
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    if args.side_integrity_guard:
        os.environ["SIDE_INTEGRITY_GUARD"] = "1"

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.outdir or f"outputs/forensics/topx_audit_{ts}")
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "per_bet_details").mkdir(parents=True, exist_ok=True)

    best_bets = _load_best_bets(Path(args.source_xlsx))
    markets = args.markets.split(",") if args.markets else None
    filtered = _apply_filters(best_bets, markets, args.min_ev)
    top_df = _select_top_x(filtered, args.top_x)
    if top_df.empty:
        raise SystemExit("No rows available after filters.")

    top_df["bet_key"] = top_df.apply(_build_bet_key, axis=1)
    top_df["odds_decimal"] = top_df["Odds"].apply(_odds_decimal_from_american)
    top_df["implied_prob_calc"] = top_df["odds_decimal"].apply(
        lambda x: 1.0 / x if pd.notna(x) and x and x > 0 else np.nan
    )

    prob_snapshot_df = None
    prob_path = Path("outputs/projections/SingleGamePropProbabilities.csv")
    prob_snapshot_file = str(prob_path.resolve()) if prob_path.exists() else None
    if prob_snapshot_file:
        prob_snapshot_df = pd.read_csv(prob_path)
        prob_snapshot_df["player_key"] = prob_snapshot_df["Player"].apply(normalize_name)
        prob_snapshot_df["game_date"] = pd.to_datetime(prob_snapshot_df["Date"], errors="coerce").dt.date

    con = _init_con(args.duckdb_path)
    try:
        players_df = _load_players(con)
        odds_df = _load_odds(con)
    finally:
        con.close()

    base_proj_path = Path("outputs/projections/BaseSingleGameProjections.csv")
    base_proj_file = str(base_proj_path.resolve()) if base_proj_path.exists() else None
    base_proj_df = _load_base_projections(base_proj_path)
    side_lookup = build_odds_side_lookup(odds_df) if not odds_df.empty else {}
    trace_df = _build_trace_keys(top_df)

    combined_rows = []
    detail_payloads = []

    proj_con = _init_con(args.duckdb_path)
    for _, row in trace_df.iterrows():
        projection = _build_projection_trace(row, proj_con, players_df, base_proj_df, prob_snapshot_df)
        odds_audit = _run_odds_audit(row, odds_df, side_lookup, args.odds_asof_join)

        odds_selected = odds_audit.get("odds_selected") or {}
        classification = _classify_row(projection, odds_audit)

        p_used = projection.get("p_used")
        odds_decimal = row.get("odds_decimal")
        ev_recomputed = (p_used * odds_decimal - 1.0) if pd.notna(p_used) and pd.notna(odds_decimal) else np.nan

        recent_stats = projection.get("recent_stat_values") or []
        recent_avg = _recent_stats_average(recent_stats)
        base_proj = projection.get("base_projection") or {}
        recent_cols = {col: np.nan for col in RECENT_GAME_COLUMNS}
        for idx, entry in enumerate(recent_stats[:RECENT_GAME_LIMIT]):
            prefix = f"recentgame{idx + 1}"
            recent_cols[f"{prefix}_game_id"] = entry.get("game_id")
            recent_cols[f"{prefix}_date"] = entry.get("game_date")
            recent_cols[f"{prefix}_value"] = entry.get("value")

        multipliers = projection.get("multipliers") or {}
        nbinom_inputs = _nbinom_inputs(row.get("Market"), row.get("Line"), projection.get("mu_used"))
        
        # Determine base_mu_input for formulas
        market_key = str(row.get("Market", "")).upper()
        base_field_map = {
            "GOALS": "mu_base_goals",
            "ASSISTS": "Assists Per Game",
            "POINTS": "Points Per Game",
            "SOG": "SOG Per Game",
            "BLOCKS": "Blocks Per Game",
        }
        base_field = base_field_map.get(market_key)
        base_mu_input = base_proj.get(base_field) if base_field else np.nan

        # Special handling for SOG using Corsi-Split logic
        base_corsi = base_proj.get("corsi_per_60_L20")
        base_thru = base_proj.get("thru_pct_L40")
        if market_key == "SOG" and pd.notna(base_corsi) and pd.notna(base_thru) and pd.notna(base_proj.get("TOI")):
            # Base Mu = (Corsi * Thru) * (BaseTOI / 60)
            # This represents the expected SOG at the *base* TOI (L10 average).
            base_mu_input = float(base_corsi) * float(base_thru) * (float(base_proj.get("TOI")) / 60.0)

        row_data = {
            "bet_key": row.get("bet_key"),
            "player": row.get("Player"),
            "market": row.get("Market"),
            "line": row.get("Line"),
            "side": row.get("Side"),
            "book": row.get("Book"),
            "odds_decimal": odds_decimal,
            "odds_american": row.get("Odds"),
            "implied_prob": row.get("Implied_Prob"),
            "implied_prob_calc": row.get("implied_prob_calc"),
            "p_used": p_used,
            "p_over_selected": projection.get("p_over_prod"),
            "p_over_raw": projection.get("p_over_raw"),
            "p_over_calibrated": projection.get("p_over_calibrated"),
            **nbinom_inputs,
            "mu_used": projection.get("mu_used"),
            "base_mu_input": base_mu_input,
            "base_corsi": base_corsi,
            "base_thru": base_thru,
            "mult_opp_sog": multipliers.get("mult_opp_sog"),
            "mult_opp_g": multipliers.get("mult_opp_g"),
            "mult_goalie": multipliers.get("mult_goalie"),
            "mult_itt": multipliers.get("mult_itt"),
            "mult_b2b": multipliers.get("mult_b2b"),
            "toi_factor": multipliers.get("toi_factor"),
            "ev_recomputed": ev_recomputed,
            "projection_status": projection.get("projection_status"),
            "projection_note": projection.get("projection_note"),
            "LINE_MATCH": projection.get("LINE_MATCH"),
            "PROB_MATH_MATCH": projection.get("PROB_MATH_MATCH"),
            "MULTIPLIER_OUTLIER": projection.get("MULTIPLIER_OUTLIER"),
            "MU_IMPLAUSIBLE": projection.get("MU_IMPLAUSIBLE"),
            "CALIBRATION_PLATEAU_EFFECT": projection.get("CALIBRATION_PLATEAU_EFFECT"),
            "ODDS_MATCHED": odds_audit.get("ODDS_MATCHED"),
            "SIDE_MATCH": odds_audit.get("SIDE_MATCH"),
            "AMBIGUOUS_MATCH_COUNT": odds_audit.get("AMBIGUOUS_MATCH_COUNT"),
            "ODDS_TIMESTAMP": odds_selected.get("odds_capture_ts_utc"),
            "SNAPSHOT_TIMESTAMP": row.get("prob_snapshot_ts"),
            "event_id_vendor": row.get("event_id_vendor"),
            "canonical_game_id": row.get("canonical_game_id"),
            "event_start_time_utc": row.get("event_start_time_utc"),
            "capture_ts_utc": row.get("capture_ts_utc"),
            "source_vendor": row.get("source_vendor"),
            "classification": classification,
            "snapshot_game_id": projection.get("snapshot_game_id"),
            "recent_stat_label": projection.get("recent_stat_label"),
            "recent_stat_avg": recent_avg,
            "prob_snapshot_file": prob_snapshot_file,
        }
        row_data.update(recent_cols)
        row_data.update(
            {
                "base_gp": base_proj.get("GP"),
                "base_toi": base_proj.get("TOI"),
                "base_sog_per_game": base_proj.get("SOG Per Game"),
                "base_goals_per_game": base_proj.get("Goals Per Game"),
                "base_assists_per_game": base_proj.get("Assists Per Game"),
                "base_points_per_game": base_proj.get("Points Per Game"),
                "base_blocks_per_game": base_proj.get("Blocks Per Game"),
                "base_mu": base_proj.get("mu_base_goals"),
                "base_projection_file": base_proj_file,
            }
        )
        combined_rows.append(row_data)

        detail_payloads.append(
            {
                "row": row,
                "projection": projection,
                "odds_audit": odds_audit,
                "classification": classification,
                "base_projection_file": base_proj_file,
                "prob_snapshot_file": prob_snapshot_file,
            }
        )
    proj_con.close()

    combined_df = pd.DataFrame(combined_rows)
    combined_csv = out_dir / "combined_summary.csv"
    combined_df.to_csv(combined_csv, index=False)
    formula_xlsx = out_dir / "combined_summary_with_formulas.xlsx"
    _write_formula_workbook(formula_xlsx, combined_df.copy())

    counts_df = combined_df["classification"].value_counts().reset_index()
    counts_df.columns = ["classification", "n"]

    combined_md = out_dir / "combined_summary.md"
    with combined_md.open("w", encoding="utf-8") as f:
        f.write("# Top-X Forensic Audit Summary\n\n")
        f.write(f"- source_xlsx: `{args.source_xlsx}`\n")
        f.write(f"- top_x: `{args.top_x}`\n")
        f.write(f"- markets_filter: `{args.markets or ''}`\n")
        f.write(f"- min_ev: `{args.min_ev}`\n")
        f.write(f"- side_integrity_guard: `{bool(args.side_integrity_guard)}`\n")
        f.write(f"- odds_asof_join: `{bool(args.odds_asof_join)}`\n\n")
        f.write("## Counts by Classification\n\n")
        f.write(counts_df.to_markdown(index=False))
        f.write("\n\n")
        f.write("## Combined Summary (Top 20)\n\n")
        f.write(_format_markdown_table(combined_df))

    market_groups = combined_df.groupby("market")
    for market, subset in market_groups:
        safe = re.sub(r"[^a-zA-Z0-9_]", "_", str(market).lower())
        market_csv = out_dir / f"combined_summary_{safe}.csv"
        subset.to_csv(market_csv, index=False)
        market_counts = subset["classification"].value_counts().reset_index()
        market_counts.columns = ["classification", "n"]
        market_md = out_dir / f"combined_summary_{safe}.md"
        with market_md.open("w", encoding="utf-8") as mf:
            mf.write(f"# Top-X Forensic Audit ({market})\n\n")
            mf.write(f"- source_xlsx: `{args.source_xlsx}`\n")
            mf.write(f"- market: `{market}`\n")
            mf.write(f"- rows: `{len(subset)}`\n\n")
            mf.write("## Counts by Classification\n\n")
            mf.write(market_counts.to_markdown(index=False))
            mf.write("\n\n")
            mf.write("## Summary Table (Top 20 within market)\n\n")
            mf.write(_format_markdown_table(subset))

    for payload in detail_payloads:
        row = payload["row"]
        projection = payload["projection"]
        odds_audit = payload["odds_audit"]
        classification = payload["classification"]

        bet_key = row.get("bet_key")
        file_safe_key = "".join(c if c.isalnum() or c in "-_" else "_" for c in str(bet_key))[:120]
        out_path = out_dir / "per_bet_details" / f"{file_safe_key}.md"

        odds_candidates = odds_audit.get("odds_candidates", [])
        odds_selected = odds_audit.get("odds_selected")
        bet_side = normalize_side(row.get("Side"))
        alt_rows = []
        if odds_candidates:
            for cand in odds_candidates:
                if normalize_side(cand.get("odds_side_raw")) == bet_side:
                    alt_rows.append(cand)

        p_used = projection.get("p_used")
        alt_ev_rows = []
        if pd.notna(p_used):
            for cand in alt_rows:
                odds_dec = cand.get("odds_decimal")
                if odds_dec:
                    alt_ev_rows.append(
                        {
                            "odds_decimal": odds_dec,
                            "odds_american": cand.get("odds_american"),
                            "alt_ev": p_used * float(odds_dec) - 1.0,
                            "odds_capture_ts_utc": cand.get("odds_capture_ts_utc"),
                        }
                    )

        with out_path.open("w", encoding="utf-8") as f:
            f.write("# Bet Forensic Detail\n\n")
            f.write(f"- bet_key: `{bet_key}`\n")
            f.write(
                f"- bet: `{row.get('Player')} | {row.get('Market')} {row.get('Line')} "
                f"{row.get('Side')} | {row.get('Book')}`\n"
            )
            f.write(f"- classification: `{classification}`\n\n")

            f.write("## What Failed\n\n")
            failed = []
            if not projection.get("PROB_MATH_MATCH", True):
                failed.append("PROB_MATH_MATCH=false")
            if projection.get("MULTIPLIER_OUTLIER"):
                failed.append("MULTIPLIER_OUTLIER=true")
            if projection.get("MU_IMPLAUSIBLE"):
                failed.append("MU_IMPLAUSIBLE=true")
            if projection.get("CALIBRATION_PLATEAU_EFFECT"):
                failed.append("CALIBRATION_PLATEAU_EFFECT=true")
            if not odds_audit.get("ODDS_MATCHED", True):
                failed.append("ODDS_MATCHED=false")
            if odds_audit.get("ODDS_MATCHED", False) and not odds_audit.get("SIDE_MATCH", True):
                failed.append("SIDE_MATCH=false")
            if odds_audit.get("AMBIGUOUS_MATCH_COUNT", 0) > 0:
                failed.append("AMBIGUOUS_MATCH_COUNT>0")
            if failed:
                for item in failed:
                    f.write(f"- {item}\n")
            else:
                f.write("- None\n")

            f.write("\n## EV Decomposition\n\n")
            f.write(f"- odds_decimal: `{row.get('odds_decimal')}`\n")
            f.write(f"- odds_american: `{row.get('Odds')}`\n")
            f.write(f"- implied_prob: `{row.get('Implied_Prob')}`\n")
            f.write(f"- p_used: `{projection.get('p_used')}`\n")
            f.write(f"- ev_recomputed: `{projection.get('ev_calc')}`\n")

            f.write("\n## Base Projection Stats\n\n")
            base_proj = projection.get("base_projection") or {}
            if base_proj:
                for key in [
                    "GP",
                    "TOI",
                    "SOG Per Game",
                    "Goals Per Game",
                    "Assists Per Game",
                    "Points Per Game",
                    "Blocks Per Game",
                ]:
                    if key in base_proj:
                        f.write(f"- {key}: `{base_proj[key]}`\n")
                if payload.get("base_projection_file"):
                    f.write(f"- source_file: `{payload.get('base_projection_file')}`\n")
            else:
                f.write("- No base projection row available for this player/team.\n")

            stat_label = projection.get("recent_stat_label")
            recent_stats = projection.get("recent_stat_values") or []
            if stat_label and recent_stats:
                f.write(f"\n## Recent Game Log ({stat_label})\n\n")
            for entry in recent_stats:
                game_id = entry.get("game_id")
                id_suffix = f" (game_id {game_id})" if game_id is not None else ""
                f.write(f"- {entry['game_date']}{id_suffix}: `{entry['value']}`\n")
            else:
                f.write("\n## Recent Game Log\n\n")
                f.write("- No recent game log values available for this market.\n")

            f.write("\n## Event Identifiers\n\n")
            for key in [
                ("canonical_game_id", projection.get("canonical_game_id")),
                ("event_id_vendor", projection.get("event_id_vendor")),
                ("snapshot_game_id", projection.get("snapshot_game_id")),
                ("event_start_time_utc", projection.get("event_start_time_utc")),
                ("capture_ts_utc", projection.get("source_capture_ts")),
                ("source_vendor", projection.get("source_vendor")),
            ]:
                if key[1] is not None:
                    f.write(f"- {key[0]}: `{key[1]}`\n")
            if payload.get("prob_snapshot_file"):
                f.write(f"- prob_snapshot_file: `{payload.get('prob_snapshot_file')}`\n")

            f.write("\n## Calculation Walkthrough\n\n")
            f.write(
                "- **p_used** = `1 - p_over_selected` (UNDER) or `p_over_selected` (OVER); Excel: `=IF([Side]=\"UNDER\", 1 - [p_over_selected], [p_over_selected])`.\n"
            )
            f.write(
                "- **Probability selection** mirrors policy; SOG uses `p_over_raw`, so Excel can use `=IF([Market]=\"ASSISTS\", [p_over_calibrated], [p_over_raw])` before storing `p_over_selected`.\n"
            )
            f.write(
                "- **mu_used** is `BaseMu * mult_opp_sog * mult_opp_g * mult_goalie * mult_itt * mult_b2b * toi_factor`; Excel formula: `=[BaseMu]*[mult_opp_sog]*[mult_opp_g]*[mult_goalie]*[mult_itt]*[mult_b2b]*[toi_factor]`.\n"
            )
            f.write(
                "- **Base SOG (2.7)** comes from `BaseSingleGameProjections.csv`; the last ten SOG rows average to that value, e.g. `=AVERAGE([recentgame1_value]:[recentgame10_value])`.\n"
            )
            f.write(
                "- **Odds EV** is `=(p_used * odds_decimal) - 1`; the CSV exposes both inputs for quick Excel verification.\n"
            )

            f.write("\n## Projection Trace Excerpt\n\n")
            f.write(f"- projection_status: `{projection.get('projection_status')}`\n")
            if projection.get("projection_note"):
                f.write(f"- projection_note: `{projection.get('projection_note')}`\n")
            f.write(f"- game_date: `{projection.get('game_date')}`\n")
            f.write(f"- mu_used: `{projection.get('mu_used')}`\n")
            f.write(f"- p_over_raw: `{projection.get('p_over_raw')}`\n")
            f.write(f"- p_over_calibrated: `{projection.get('p_over_calibrated')}`\n")
            f.write(f"- p_over_selected: `{projection.get('p_over_prod')}`\n")
            f.write(f"- LINE_MATCH: `{projection.get('LINE_MATCH')}`\n")
            f.write(f"- PROB_MATH_MATCH: `{projection.get('PROB_MATH_MATCH')}`\n")
            f.write(f"- CALIBRATION_PLATEAU_EFFECT: `{projection.get('CALIBRATION_PLATEAU_EFFECT')}`\n")
            f.write(f"- MULTIPLIER_OUTLIER: `{projection.get('MULTIPLIER_OUTLIER')}`\n")
            f.write(f"- MU_IMPLAUSIBLE: `{projection.get('MU_IMPLAUSIBLE')}`\n")
            for key, val in (projection.get("multipliers") or {}).items():
                f.write(f"- {key}: `{val}`\n")

            f.write("\n## Odds Join Excerpt\n\n")
            if odds_selected:
                f.write("### Selected Odds Row\n\n")
                for key in [
                    "odds_capture_ts_utc",
                    "odds_event_date",
                    "odds_market_raw",
                    "odds_line",
                    "odds_side_raw",
                    "odds_side_interpreted",
                    "odds_american",
                    "odds_decimal",
                    "odds_book_raw",
                ]:
                    f.write(f"- {key}: `{odds_selected.get(key)}`\n")
                f.write(f"- candidate_count: `{odds_selected.get('candidate_count')}`\n")
            else:
                f.write("No odds candidates matched.\n")

            f.write("\n### Candidate Odds Rows (up to 5)\n\n")
            if odds_candidates:
                odds_df = pd.DataFrame(odds_candidates).head(5)
                f.write(odds_df.to_markdown(index=False))
                f.write("\n")
            else:
                f.write("No candidate rows.\n")

            f.write("\n## Diff Lens\n\n")
            if classification == "ODDS JOIN / SIDE ISSUE":
                if alt_ev_rows:
                    f.write("### Alternate Odds EV (matching bet side)\n\n")
                    f.write(pd.DataFrame(alt_ev_rows).to_markdown(index=False))
                    f.write("\n")
                else:
                    f.write("No alternate odds rows matching bet side.\n")
            elif classification == "CALIBRATION_PLATEAU_EFFECT":
                k = int(np.floor(float(row.get("Line"))) + 1) if pd.notna(row.get("Line")) else ""
                f.write(f"- plateau_id: `{row.get('Market')}_k{k}`\n")
                f.write("- bucket_size: `1` (assumed for line -> k mapping)\n")
                f.write(f"- p_over_raw: `{projection.get('p_over_raw')}`\n")
                f.write(f"- p_over_calibrated: `{projection.get('p_over_calibrated')}`\n")
            elif classification == "PROJECTION_INPUT/MULTIPLIER ISSUE":
                triggers = []
                if not projection.get("PROB_MATH_MATCH", True):
                    triggers.append("probability math mismatch")
                if projection.get("MULTIPLIER_OUTLIER"):
                    triggers.append("multiplier outlier")
                if projection.get("MU_IMPLAUSIBLE"):
                    triggers.append("mu implausible")
                f.write(f"- triggers: `{', '.join(triggers)}`\n")
            else:
                f.write("No diff lens issues; bet passes all checks.\n")

    readme = out_dir / "README.md"
    with readme.open("w", encoding="utf-8") as f:
        f.write("# Top-X Forensic Audit Output\n\n")
        f.write("## Overview\n")
        f.write(
            "This folder contains a combined projection trace + odds side-integrity audit for the "
            "Top-X bets from the best-bets workbook. Diagnostics are read-only and do not modify "
            "best-bets outputs.\n\n"
        )
        f.write("## Artifacts\n")
        f.write("- combined_summary.csv: One row per bet with projection + odds flags and classification.\n")
        f.write("- combined_summary.md: Narrative summary with counts by classification.\n")
        f.write("- combined_summary_with_formulas.xlsx: Excel workbook that writes the recomputed formulas (`p_used`, `p_over_selected`, `p_over_raw`, `p_over_calibrated`, `mu_used`, etc.) plus diff columns so you can see their calculations live.\n")
        f.write("- combined_summary_<market>.csv/.md: Market-specific slices so you can inspect Goals/Assists/Points/Blocks/SOG individually.\n")
        f.write("- per_bet_details/: One markdown file per bet with trace + odds excerpts.\n")
        f.write("- README.md: This guide.\n\n")
        f.write("## How to rerun\n")
        f.write(
            "python scripts/forensics/run_topx_forensic_audit.py "
            "--top-x 50 "
            "--source-xlsx outputs/ev_analysis/ev_bets_ranked.xlsx\n\n"
        )
        f.write("## Flags\n")
        f.write("- --markets: Filter markets (comma-separated), applied before Top-X ranking.\n")
        f.write("- --min-ev: Minimum EV% threshold, applied before Top-X ranking.\n")
        f.write("- --side-integrity-guard: Sets SIDE_INTEGRITY_GUARD=1 for this run only.\n")
        f.write("- --odds-asof-join: Prefer odds rows captured at/earlier than bet capture_ts_utc.\n\n")
        f.write("## Assumptions & Diagnostics Notes\n")
        f.write(
            "- game_date inferred from event_start_time_utc, else prob_snapshot_ts.\n"
            "- player_id resolved by normalized name, team used if available.\n"
            "- projection fallback uses outputs/projections/SingleGamePropProbabilities.csv when DB snapshot is missing.\n"
            "- if no snapshot data is available, best-bets projection fields are used and marked OK_BEST_BETS.\n"
            "- Base projections from outputs/projections/BaseSingleGameProjections.csv provide the SOG/goal/assist/point/block averages and feed the underline prob math.\n"
            "- Recent game logs are queried from fact_skater_game_situation for the last 10 contests (per-market stat: goals, assists, points, SOG, blocks).\n"
            "- PROB_MATH_MATCH compares recomputed p_over_* against best-bets fields within 1e-6.\n"
            "- CALIBRATION_PLATEAU_EFFECT flags ASSISTS/POINTS line 0.5 when calibrated != raw.\n"
            "- MULTIPLIER_OUTLIER flags multipliers outside [0.5, 1.5].\n"
            "- MU_IMPLAUSIBLE flags mu <= 0 or mu > 6.\n"
            "- ODDS_MATCHED and SIDE_MATCH are based on fact_prop_odds join using normalized keys.\n"
            "- odds-asof join uses capture_ts_utc if present; otherwise falls back to standard join.\n"
            "- plateau_id uses market + k (line->k mapping). No explicit bucket metadata available.\n"
        )
        f.write(
            "- Combined summary rows now surface `recent_stat_label`, `recent_stat_values`, `recent_stat_avg`, `snapshot_game_id`, `base_gp`, `base_toi`, `base_sog_per_game`, "
            "`base_goals_per_game`, `base_assists_per_game`, `base_points_per_game`, `base_blocks_per_game`, `prob_snapshot_file`, and `base_projection_file` so the spreadsheet documents base stats and identifiers.\n"
        )
        f.write(
            "- Each row now emits `recentgame1_game_id/date/value` through `recentgame10_game_id/date/value`, so the last ten `fact_skater_game_situation` entries (Goals/Assists/Points/SOG/Blocks, depending on the market) are captured explicitly for your trace and auditing needs.\n"
        )
        f.write(
            "- The combined summary columns are now ordered to keep the bet/odds info together, followed by the full model trace (`p_used`, `p_over_selected`, `p_over_raw`, `p_over_calibrated`, `mu_used`, multipliers) before the base stats and recent games, helping you step through the maths from probability back to the historical averages.\n"
        )
        f.write(
            "- `nbinom_alpha`, `nbinom_k`, `nbinom_n`, and `nbinom_p` give you the inputs for the Excel formula (`1 - NBINOM.DIST(k-1, n, p, TRUE)`) that reconstructs the SOG/Blocks raw probability; the `p_over_raw_formula` column now references those values instead of simply mirroring the stored probability.\n"
        )
        f.write(
            "- The formula workbook now surfaces `p_over_raw_formula`, `p_over_calibrated_formula`, and their `*_delta` columns so you can confirm the stored probabilities match the recomputed values.\n"
        )
        f.write(
            "- `snapshot_game_id` exposes the NHL API Game ID for the projection snapshot; the CSV references (`prob_snapshot_file`, `base_projection_file`) show exactly which files provided the SOG/Goals/Assists/Points/Blocks averages.\n"
        )

    print(f"Wrote combined summary to {combined_csv}")
    print(f"Wrote combined markdown to {combined_md}")
    print(f"Wrote per-bet details to {out_dir / 'per_bet_details'}")
    print(f"Wrote README to {readme}")


if __name__ == "__main__":
    main()
